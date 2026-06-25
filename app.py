# ============================================================
#  APP STREAMLIT - CONSOLIDADOR DE SUMAS Y SALDOS
#  Año N vs Año N-1
# ============================================================

import io
import re
import traceback
import numbers

import pandas as pd
import streamlit as st
import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string


# ============================================================
# CONFIGURACIÓN DE STREAMLIT
# ============================================================

st.set_page_config(
    page_title="Consolidador Sumas y Saldos",
    page_icon="📊",
    layout="wide",
)

st.title("📊 Consolidador de Sumas y Saldos")
st.caption("Consolida el año N y el año N-1 en un nuevo Excel, con el bloque del año N-1 desplazado desde la columna AE.")


# ============================================================
# EXCEPCIONES
# ============================================================

class AppValidationError(Exception):
    """Error controlado para mostrar mensajes claros en la interfaz."""
    pass


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def normalizar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def limpiar_nombre_columna(columna) -> str:
    texto = str(columna).strip().replace("\n", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def get_excel_sheet_names(file_bytes: bytes) -> list[str]:
    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        return xls.sheet_names
    except Exception as e:
        raise AppValidationError(f"No se pudo leer el archivo Excel. Verifica que sea un .xlsx válido. Detalle: {e}")


def seleccionar_hoja_por_defecto(sheet_names: list[str], preferida: str = "Holded") -> int:
    if preferida in sheet_names:
        return sheet_names.index(preferida)
    return 0


def convertir_numero(valor):
    """
    Convierte valores numéricos que pueden venir como texto.

    Soporta:
    - 1234.56
    - 1,234.56
    - 1.234,56
    - 1234,56
    - -1.234,56
    - 1.234,56-
    - (1.234,56)
    - € 1.234,56
    """
    if pd.isna(valor):
        return None

    if isinstance(valor, numbers.Number) and not isinstance(valor, bool):
        return float(valor)

    texto = str(valor).strip()

    if texto == "" or texto in ["-", "—"]:
        return None

    texto = (
        texto.replace("€", "")
        .replace(" ", "")
        .replace("\u00a0", "")
        .replace("−", "-")
    )

    negativo = False

    # Formato contable: (1.234,56)
    if texto.startswith("(") and texto.endswith(")"):
        negativo = True
        texto = texto[1:-1]

    # Negativo al final: 1.234,56-
    if texto.endswith("-"):
        negativo = True
        texto = texto[:-1]

    # Negativo al inicio: -1.234,56
    if texto.startswith("-"):
        negativo = True
        texto = texto[1:]

    # Caso europeo: 1.234,56
    if "." in texto and "," in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")

    # Caso decimal con coma: 1234,56
    elif "," in texto and "." not in texto:
        texto = texto.replace(",", ".")

    try:
        numero = float(texto)
        return -numero if negativo else numero
    except Exception:
        return None


def numero_o_cero(valor) -> float:
    numero = convertir_numero(valor)
    if numero is None:
        return 0.0
    return float(numero)


def convertir_cuenta(valor, cuenta_como_numero: bool):
    """
    Convierte la cuenta contable a número si el usuario lo activa.
    Si no puede convertirse limpiamente, se deja como texto.
    """
    texto = normalizar_texto(valor)

    if texto == "":
        return ""

    if not cuenta_como_numero:
        return texto

    limpio = texto.replace(" ", "").replace("\u00a0", "")

    # Corrige cuentas que llegan como 430000.0
    if re.fullmatch(r"\d+\.0+", limpio):
        limpio = limpio.split(".")[0]

    # Solo convierte si es una cuenta puramente numérica.
    # Así evitamos convertir cuentas con guiones, puntos internos o letras.
    if re.fullmatch(r"\d+", limpio):
        try:
            return int(limpio)
        except Exception:
            return texto

    return texto


def es_columna_texto(nombre_columna: str) -> bool:
    nombre = nombre_columna.strip().lower()

    palabras_texto = [
        "nombre",
        "descripción",
        "descripcion",
        "concepto",
        "detalle",
        "texto",
        "observación",
        "observacion",
    ]

    return any(palabra in nombre for palabra in palabras_texto)


def convertir_columnas_numericas(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convierte automáticamente columnas que parecen numéricas.
    No convierte Cuenta aquí, porque Cuenta tiene su propia lógica.
    """
    for col in df.columns:
        if col == "Cuenta":
            continue

        if es_columna_texto(col):
            continue

        serie_original = df[col]
        valores_no_vacios = serie_original.dropna().astype(str).str.strip()
        valores_no_vacios = valores_no_vacios[valores_no_vacios != ""]

        if len(valores_no_vacios) == 0:
            continue

        serie_convertida = serie_original.apply(convertir_numero)
        cantidad_numericos = serie_convertida.notna().sum()

        # Si la mayoría de valores parecen números, convertimos la columna completa.
        if cantidad_numericos >= len(valores_no_vacios) * 0.60:
            df[col] = serie_original.apply(numero_o_cero)

    return df


def fila_contiene_texto_eliminar(fila: pd.Series, textos_a_eliminar: list[str]) -> bool:
    texto_fila = " ".join([normalizar_texto(v) for v in fila.values]).lower()

    for texto in textos_a_eliminar:
        if texto.strip() and texto.strip().lower() in texto_fila:
            return True

    return False


def limpiar_lista_textos(texto_multilinea: str) -> list[str]:
    textos = []
    for linea in texto_multilinea.splitlines():
        linea = linea.strip()
        if linea:
            textos.append(linea)
    return textos


def dataframe_preview_seguro(df: pd.DataFrame, max_rows: int = 20) -> pd.DataFrame:
    if df.empty:
        return df
    return df.head(max_rows)


def valor_excel(valor):
    """Evita escribir NaN/NaT en Excel."""
    if pd.isna(valor):
        return None

    if isinstance(valor, numbers.Number) and not isinstance(valor, bool):
        if float(valor).is_integer():
            return int(valor)
        return float(valor)

    return valor


# ============================================================
# LECTURA Y LIMPIEZA
# ============================================================

def leer_y_limpiar_excel(
    file_bytes: bytes,
    sheet_name: str,
    skip_rows: int,
    etiqueta: str,
    textos_a_eliminar: list[str],
    cuenta_como_numero: bool,
    acreedor_como_negativo: bool,
):
    """
    Lee, valida, limpia y calcula Saldo.
    Devuelve:
    - DataFrame limpio
    - Diccionario de resumen
    - Lista de advertencias
    """
    advertencias = []

    try:
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet_name,
            skiprows=skip_rows,
            header=0,
        )
    except Exception as e:
        raise AppValidationError(f"No se pudo leer la hoja '{sheet_name}' del archivo {etiqueta}. Detalle: {e}")

    filas_originales = len(df)

    if df.empty:
        raise AppValidationError(f"El archivo {etiqueta} quedó vacío al leerlo. Revisa la hoja y las filas a ignorar.")

    # Limpiar encabezados
    df.columns = [limpiar_nombre_columna(col) for col in df.columns]

    # Eliminar columnas totalmente vacías
    df = df.dropna(axis=1, how="all").copy()

    columnas_requeridas = ["Cuenta", "Saldo deudor", "Saldo acreedor"]
    faltantes = [col for col in columnas_requeridas if col not in df.columns]

    if faltantes:
        raise AppValidationError(
            f"En {etiqueta} faltan columnas obligatorias: {', '.join(faltantes)}. "
            f"Columnas encontradas: {', '.join(map(str, df.columns))}"
        )

    # Eliminar filas totalmente vacías
    df = df.dropna(how="all").copy()

    # Eliminar filas sin Cuenta
    df = df[df["Cuenta"].notna()].copy()
    df["Cuenta"] = df["Cuenta"].apply(normalizar_texto)
    df = df[df["Cuenta"] != ""].copy()

    # Eliminar Total / Informe creado automáticamente / textos configurados
    mascara_invalida = df.apply(
        lambda fila: fila_contiene_texto_eliminar(fila, textos_a_eliminar),
        axis=1
    )

    filas_eliminadas_texto = int(mascara_invalida.sum())
    df = df[~mascara_invalida].copy()

    if df.empty:
        raise AppValidationError(f"El archivo {etiqueta} quedó sin filas válidas después de limpiar.")

    # Convertir Cuenta
    cuentas_antes = df["Cuenta"].copy()
    df["Cuenta"] = df["Cuenta"].apply(lambda x: convertir_cuenta(x, cuenta_como_numero))

    if cuenta_como_numero:
        cuentas_no_convertidas = 0
        for original, convertido in zip(cuentas_antes, df["Cuenta"]):
            if normalizar_texto(original) != "" and isinstance(convertido, str):
                cuentas_no_convertidas += 1

        if cuentas_no_convertidas > 0:
            advertencias.append(
                f"{etiqueta}: {cuentas_no_convertidas} cuentas no pudieron convertirse a número "
                f"y se conservaron como texto."
            )

    # Convertir columnas numéricas generales
    df = convertir_columnas_numericas(df)

    # Convertir saldos obligatorios
    df["Saldo deudor"] = df["Saldo deudor"].apply(numero_o_cero)
    df["Saldo acreedor"] = df["Saldo acreedor"].apply(numero_o_cero)

    if acreedor_como_negativo:
        # Si viene positivo, se vuelve negativo.
        # Si ya viene negativo, se conserva negativo.
        df["Saldo acreedor"] = df["Saldo acreedor"].apply(lambda x: -abs(x) if x != 0 else 0.0)

    df["Saldo"] = df["Saldo deudor"] + df["Saldo acreedor"]

    resumen = {
        "archivo": etiqueta,
        "filas_originales": filas_originales,
        "filas_validas": len(df),
        "filas_eliminadas_texto": filas_eliminadas_texto,
        "columnas": len(df.columns),
        "saldo_total": float(df["Saldo"].sum()),
    }

    return df.reset_index(drop=True), resumen, advertencias


# ============================================================
# VALIDACIONES
# ============================================================

def validar_compatibilidad(df_n: pd.DataFrame, df_n1: pd.DataFrame) -> list[str]:
    advertencias = []

    cols_n = set(df_n.columns) - {"Saldo"}
    cols_n1 = set(df_n1.columns) - {"Saldo"}

    if cols_n != cols_n1:
        solo_en_n = sorted(list(cols_n - cols_n1))
        solo_en_n1 = sorted(list(cols_n1 - cols_n))

        if solo_en_n:
            advertencias.append(f"Columnas solo en Año N: {', '.join(solo_en_n)}")

        if solo_en_n1:
            advertencias.append(f"Columnas solo en Año N-1: {', '.join(solo_en_n1)}")

    return advertencias


def validar_saldo(df: pd.DataFrame, etiqueta: str):
    diff = (df["Saldo"] - (df["Saldo deudor"] + df["Saldo acreedor"])).abs().max()

    if pd.isna(diff):
        diff = 0

    if diff > 0.0001:
        raise AppValidationError(
            f"La columna Saldo de {etiqueta} no coincide con Saldo deudor + Saldo acreedor. "
            f"Diferencia máxima: {diff:.4f}"
        )


# ============================================================
# GENERACIÓN DEL EXCEL
# ============================================================

def aplicar_formato_celda(
    cell,
    col_name,
    valor,
    cuenta_como_numero,
    font_data,
    font_cuenta,
    align_left,
    align_right,
    border_data,
    num_fmt_num,
):
    cell.border = border_data

    if col_name == "Cuenta":
        cell.font = font_cuenta

        if cuenta_como_numero and isinstance(valor, numbers.Number):
            cell.alignment = align_right
            cell.number_format = "0"
        else:
            cell.alignment = align_left
            cell.number_format = "General"

    elif es_columna_texto(col_name):
        cell.font = font_data
        cell.alignment = align_left
        cell.number_format = "General"

    elif isinstance(valor, numbers.Number) and not isinstance(valor, bool):
        cell.font = font_data
        cell.alignment = align_right
        cell.number_format = num_fmt_num

    else:
        cell.font = font_data
        cell.alignment = align_left
        cell.number_format = "General"


def generar_excel_consolidado(
    df_n: pd.DataFrame,
    df_n1: pd.DataFrame,
    cuenta_como_numero: bool,
    columna_inicio_n1: str = "AD",
) -> bytes:
    """
    Crea el Excel consolidado en memoria.

    Estructura:
    - Año N arriba desde columna A.
    - Año N-1 debajo, sin fila en blanco.
    - Año N-1:
        Columna A = Cuenta.
        Desde columna AD = resto de columnas, excluyendo Cuenta.
    - No se repiten encabezados del Año N-1.
    """
    offset_n1 = column_index_from_string(columna_inicio_n1.upper())

    cols_n = list(df_n.columns)
    cols_n1 = list(df_n1.columns)

    cols_n1_sin_cuenta = [col for col in cols_n1 if col != "Cuenta"]

    wb = Workbook()
    ws = wb.active
    ws.title = "SyS Holded"

    # Estilos
    fill_title = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
    fill_header = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")

    font_title = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    font_header = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    font_data = Font(name="Arial", size=10)
    font_cuenta = Font(name="Arial", size=10, bold=True)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin = Side(border_style="thin", color="CCCCCC")
    border_data = Border(bottom=thin)

    num_fmt_num = '#,##0.00;[Red]-#,##0.00;0.00'

    ultima_columna = max(
        len(cols_n),
        offset_n1 + len(cols_n1_sin_cuenta) - 1
    )

    # Título
    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=ultima_columna,
    )

    titulo = ws.cell(row=1, column=1, value="SUMAS Y SALDOS CONSOLIDADO — AÑO N / AÑO N-1")
    titulo.font = font_title
    titulo.fill = fill_title
    titulo.alignment = align_center
    ws.row_dimensions[1].height = 24

    # Encabezados Año N
    header_row_n = 2

    for col_idx, col_name in enumerate(cols_n, start=1):
        cell = ws.cell(row=header_row_n, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    ws.row_dimensions[header_row_n].height = 20

    # Datos Año N
    first_data_row_n = header_row_n + 1

    for row_idx, (_, fila) in enumerate(df_n.iterrows(), start=first_data_row_n):
        for col_idx, col_name in enumerate(cols_n, start=1):
            valor = valor_excel(fila[col_name])

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)

            aplicar_formato_celda(
                cell=cell,
                col_name=col_name,
                valor=valor,
                cuenta_como_numero=cuenta_como_numero,
                font_data=font_data,
                font_cuenta=font_cuenta,
                align_left=align_left,
                align_right=align_right,
                border_data=border_data,
                num_fmt_num=num_fmt_num,
            )

    last_data_row_n = first_data_row_n + len(df_n) - 1

    # Año N-1 empieza inmediatamente debajo del Año N
    first_data_row_n1 = last_data_row_n + 1

    for row_idx, (_, fila) in enumerate(df_n1.iterrows(), start=first_data_row_n1):

        # Columna A = Cuenta del Año N-1
        valor_cuenta = valor_excel(fila.get("Cuenta", ""))

        cell_cuenta = ws.cell(row=row_idx, column=1, value=valor_cuenta)

        aplicar_formato_celda(
            cell=cell_cuenta,
            col_name="Cuenta",
            valor=valor_cuenta,
            cuenta_como_numero=cuenta_como_numero,
            font_data=font_data,
            font_cuenta=font_cuenta,
            align_left=align_left,
            align_right=align_right,
            border_data=border_data,
            num_fmt_num=num_fmt_num,
        )

        # Resto de columnas del Año N-1 desde AD, excluyendo Cuenta
        for col_idx, col_name in enumerate(cols_n1_sin_cuenta, start=offset_n1):
            valor = valor_excel(fila[col_name])

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)

            aplicar_formato_celda(
                cell=cell,
                col_name=col_name,
                valor=valor,
                cuenta_como_numero=cuenta_como_numero,
                font_data=font_data,
                font_cuenta=font_cuenta,
                align_left=align_left,
                align_right=align_right,
                border_data=border_data,
                num_fmt_num=num_fmt_num,
            )

    # Ajuste de anchos
    for col_idx in range(1, ultima_columna + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 35)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 14)
    ws.column_dimensions[columna_inicio_n1.upper()].width = max(
        ws.column_dimensions[columna_inicio_n1.upper()].width,
        14
    )

    # Congelar encabezado
    ws.freeze_panes = "B3"

    # Filtro en la zona principal del Año N
    if len(cols_n) > 0 and len(df_n) > 0:
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols_n))}{last_data_row_n}"

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:
    st.header("⚙️ Configuración")

    skip_rows = st.number_input(
        "Filas iniciales a ignorar",
        min_value=0,
        max_value=50,
        value=5,
        step=1,
        help="Para Holded normalmente son 5 filas."
    )

    columna_inicio_n1 = st.text_input(
        "Columna donde empieza el resto de datos del Año N-1",
        value="AE",
        max_chars=3,
        help="Cuenta del Año N-1 irá en A. El resto de datos empezará aquí."
    ).strip().upper()

    cuenta_como_numero = st.checkbox(
        "Convertir Cuenta a número",
        value=True,
        help="Actívalo para evitar el aviso de Excel de número almacenado como texto. Desactívalo si tus cuentas tienen ceros a la izquierda."
    )

    acreedor_como_negativo = st.checkbox(
        "Tratar Saldo acreedor como negativo",
        value=True,
        help="Recomendado si el archivo trae los acreedores como positivos pero quieres saldo neto."
    )

    textos_default = "Total\nInforme creado automáticamente\nInforme creado automaticamente"

    textos_eliminar_input = st.text_area(
        "Textos de filas a eliminar",
        value=textos_default,
        height=110,
        help="Cualquier fila que contenga estos textos será eliminada."
    )

    textos_a_eliminar = limpiar_lista_textos(textos_eliminar_input)

    nombre_salida = st.text_input(
        "Nombre del archivo de salida",
        value="Consolidado_Sumas_y_Saldos.xlsx"
    )

    if not nombre_salida.lower().endswith(".xlsx"):
        nombre_salida += ".xlsx"


# ============================================================
# CARGA DE ARCHIVOS
# ============================================================

col1, col2 = st.columns(2)

with col1:
    archivo_n = st.file_uploader(
        "📘 Archivo Año N",
        type=["xlsx"],
        key="archivo_n"
    )

with col2:
    archivo_n1 = st.file_uploader(
        "📗 Archivo Año N-1",
        type=["xlsx"],
        key="archivo_n1"
    )


# ============================================================
# VALIDACIÓN DE ENTRADAS
# ============================================================

if archivo_n is None or archivo_n1 is None:
    st.info("Sube los dos archivos Excel para continuar.")
    st.stop()

try:
    archivo_n_bytes = archivo_n.getvalue()
    archivo_n1_bytes = archivo_n1.getvalue()

    sheets_n = get_excel_sheet_names(archivo_n_bytes)
    sheets_n1 = get_excel_sheet_names(archivo_n1_bytes)

except AppValidationError as e:
    st.error(str(e))
    st.stop()


col3, col4 = st.columns(2)

with col3:
    hoja_n = st.selectbox(
        "Hoja del Año N",
        options=sheets_n,
        index=seleccionar_hoja_por_defecto(sheets_n),
    )

with col4:
    hoja_n1 = st.selectbox(
        "Hoja del Año N-1",
        options=sheets_n1,
        index=seleccionar_hoja_por_defecto(sheets_n1),
    )


# ============================================================
# PROCESAMIENTO
# ============================================================

procesar = st.button("🚀 Generar consolidado", type="primary", use_container_width=True)

if procesar:
    try:
        # Validar columna de inicio
        try:
            column_index_from_string(columna_inicio_n1)
        except Exception:
            raise AppValidationError("La columna de inicio para Año N-1 no es válida. Usa algo como AE, AF, AG, etc.")

        with st.status("Procesando archivos...", expanded=True) as status:
            st.write("Leyendo y limpiando Año N...")
            df_n, resumen_n, advertencias_n = leer_y_limpiar_excel(
                file_bytes=archivo_n_bytes,
                sheet_name=hoja_n,
                skip_rows=skip_rows,
                etiqueta="Año N",
                textos_a_eliminar=textos_a_eliminar,
                cuenta_como_numero=cuenta_como_numero,
                acreedor_como_negativo=acreedor_como_negativo,
            )

            st.write("Leyendo y limpiando Año N-1...")
            df_n1, resumen_n1, advertencias_n1 = leer_y_limpiar_excel(
                file_bytes=archivo_n1_bytes,
                sheet_name=hoja_n1,
                skip_rows=skip_rows,
                etiqueta="Año N-1",
                textos_a_eliminar=textos_a_eliminar,
                cuenta_como_numero=cuenta_como_numero,
                acreedor_como_negativo=acreedor_como_negativo,
            )

            st.write("Validando saldos...")
            validar_saldo(df_n, "Año N")
            validar_saldo(df_n1, "Año N-1")

            advertencias_formato = validar_compatibilidad(df_n, df_n1)

            st.write("Generando Excel final...")
            excel_bytes = generar_excel_consolidado(
                df_n=df_n,
                df_n1=df_n1,
                cuenta_como_numero=cuenta_como_numero,
                columna_inicio_n1=columna_inicio_n1,
            )

            status.update(label="Consolidado generado correctamente", state="complete", expanded=False)

        # ====================================================
        # RESULTADOS
        # ====================================================

        st.success("✅ Archivo consolidado generado correctamente.")

        metric1, metric2, metric3, metric4 = st.columns(4)

        with metric1:
            st.metric("Filas Año N", resumen_n["filas_validas"])

        with metric2:
            st.metric("Filas Año N-1", resumen_n1["filas_validas"])

        with metric3:
            st.metric("Eliminadas Año N", resumen_n["filas_eliminadas_texto"])

        with metric4:
            st.metric("Eliminadas Año N-1", resumen_n1["filas_eliminadas_texto"])

        todas_advertencias = advertencias_n + advertencias_n1 + advertencias_formato

        if todas_advertencias:
            st.warning("El proceso terminó con advertencias. Revisa antes de usar el archivo final.")
            with st.expander("Ver advertencias"):
                for advertencia in todas_advertencias:
                    st.write(f"• {advertencia}")

        st.download_button(
            label="⬇️ Descargar Excel consolidado",
            data=excel_bytes,
            file_name=nombre_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        tab1, tab2 = st.tabs(["Vista previa Año N", "Vista previa Año N-1"])

        with tab1:
            st.dataframe(
                dataframe_preview_seguro(df_n),
                use_container_width=True,
                hide_index=True,
            )

        with tab2:
            st.dataframe(
                dataframe_preview_seguro(df_n1),
                use_container_width=True,
                hide_index=True,
            )

    except AppValidationError as e:
        st.error("No se pudo generar el consolidado.")
        st.write(str(e))

    except Exception as e:
        st.error("Ocurrió un error inesperado.")
        st.write(str(e))

        with st.expander("Detalle técnico del error"):
            st.code(traceback.format_exc())
